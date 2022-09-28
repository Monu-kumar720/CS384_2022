
import pandas as pd #importing pandas
import numpy as np #importing numpy
import itertools    #This is to iterate through multiple lists
import os          ##importing os
import openpyxl   ##importing openpyxl for excel
from openpyxl.styles.borders import Border, Side  ## importing some styles
from openpyxl import Workbook,load_workbook  ##from openpyxl taking workbook and load_book
wb=load_workbook('input_octant_transition_identify.xlsx')  #taking input file
ws=wb.active  #the active sheet on which i am working
df=pd.read_excel('input_octant_transition_identify.xlsx')  ##reading input file by help of pandas
## taking the mean of columns U,V,W and taking it till 5 decimal places
x=df['U'].mean() 
x=round(x,5)
ws['E1']='U Avg'
ws['E2']=x
y=df['V'].mean()
y=round(y,5)
ws['F1']='V Avg'
ws['F2']=y
z=df['W'].mean()
z=round(z,5)
ws['G1']='U Avg'
ws['G2']=z

## 3 empty list and appending each iteam value with its average value
U_avg=[]
V_avg=[]
W_avg=[]
for iteam in df['U']:
    U_avg.append(float(iteam)-x)
for iteam in df['V']:
    V_avg.append(float(iteam)-y)
for iteam in df['W']:
    W_avg.append(float(iteam)-z)
y=2 #row no 2
ws.insert_cols(8) #inserting a column at position 8
cell_title = ws.cell(row=1, column=8) #defining the cell title
cell_title.value = "U'=U-U Avg" # naming the cell title
# inserting the values of list U_avg in the column
for x in range(len(U_avg)):
    cell_to_write = ws.cell(row=y, column=8)
    cell_to_write.value = U_avg[x]
    y += 1
y=2
ws.insert_cols(9)
cell_title = ws.cell(row=1, column=9)
cell_title.value = "V'=V-V Avg"

for x in range(len(V_avg)):
    cell_to_write = ws.cell(row=y, column=9)
    cell_to_write.value = V_avg[x]
    y += 1
y=2
ws.insert_cols(10)
cell_title = ws.cell(row=1, column=10)
cell_title.value = "W'=W-W Avg"

for x in range(len(W_avg)):
    cell_to_write = ws.cell(row=y, column=10)
    cell_to_write.value = W_avg[x]
    y += 1

Octant=[]  # -->creating an empty list 
    ## --->>traversing through U_avg,V_avg,W_avg  simultaneously and storing the values in list Octant
for (a,b,c) in zip(U_avg,V_avg,W_avg):

    if (a>0 and b>0 and c>0):  # --> this tells that a,b,c is 1st octant
        Octant.append(+1)
        
    if (a>0 and b>0 and c<0): # --> this tells that a,b,c is 2nd octant
        Octant.append(-1)
        
    if (a<0 and b>0 and c>0): # --> this tells that a,b,c is 3rd octant
        Octant.append(+2)
        
    if (a<0 and b>0 and c<0): # --> this tells that a,b,c is 4th octant
        Octant.append(-2)
        
    if (a<0 and b<0 and c>0): # --> this tells that a,b,c is 5th octant
        Octant.append(+3)
            
    if (a<0 and b<0 and c<0): # --> this tells that a,b,c is 6th octant
        Octant.append(-3)
        
    if (a>0 and b<0 and c>0): # --> this tells that a,b,c is 7th octant
        Octant.append(+4)
        
    if (a>0 and b<0 and c<0): # --> this tells that a,b,c is 8th octant
        Octant.append(-4)

y=2# row no 2
ws.insert_cols(11) #inserting a column at position 11
cell_title = ws.cell(row=1, column=11) #creating a cell at row 1 and col 11
cell_title.value = "Octant"  ## naming the cell 
##inserting the elements of octant in the column
for x in range(len(Octant)):
    cell_to_write = ws.cell(row=y, column=11)
    cell_to_write.value = Octant[x]
    y += 1
##defing a list
octantID=['',1,-1,2,-2,3,-3,4,-4]  ##defing a list 

OverallCount=['OverallCount',Octant.count(1),Octant.count(-1),Octant.count(2),Octant.count(-2),Octant.count(3),Octant.count(-3),Octant.count(4),Octant.count(-4)]
###--->>>taking the mod value
Mod=5000  
y=2
x=0
## this is for making boarder in the
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
## putting values of counts of 1,-1 and so on in the given range
for i in range(13,22):
    ws.insert_cols(i)
    cell_title = ws.cell(row=1, column=i)
    ws.cell(row=1, column=i).border = thin_border
    cell_title.value = octantID[x]
    cell_to_write = ws.cell(row=y, column=i)
    ws.cell(row=y, column=i).border = thin_border
    cell_to_write.value = OverallCount[x]
    x+=1
cell_to_write = ws.cell(row=3, column=12)
cell_to_write.value = 'User Input'
cell_to_write = ws.cell(row=3, column=13)
cell_to_write.value = 'Mod '+str(Mod)

i=0 ## taking i=0
size=len(U_avg) ## length of U_avg
y=4
while i<size:
        x=0
        
        arr=[]  ##empty list
        
        if i+Mod>size:  ## in case when at the last equal division is not possible
            # creating a arr which has all counts and range
            arr=[str(i)+"-"+str(size),Octant[i:].count(1),Octant[i:].count(-1),Octant[i:].count(2),Octant[i:].count(-2),Octant[i:].count(3),Octant[i:].count(-3),Octant[i:].count(4),Octant[i:].count(-4)]
            ##--> inserting  the arr1by row in the file1
            for j in range(13,22):
                cell_to_write = ws.cell(row=y, column=j)
                ws.cell(row=y, column=j).border = thin_border
                cell_to_write.value = arr[x]
                OverallCount[0]='Verified'
                cell_to_write = ws.cell(row=y+1, column=j)
                ws.cell(row=y+1, column=j).border = thin_border
                cell_to_write.value = OverallCount[x]
                x+=1
            
            break ## breaking the while loop
            
        else:
            # creating a arr which has all counts and range
            arr=[str(i)+"-"+str(i+Mod-1),Octant[i:i+Mod].count(1),Octant[i:i+Mod].count(-1),Octant[i:i+Mod].count(2),Octant[i:i+Mod].count(-2),Octant[i:i+Mod].count(3),Octant[i:i+Mod].count(-3),Octant[i:i+Mod].count(4),Octant[i:i+Mod].count(-4)]
            ##--> inserting  the arr by row in the file1
            for j in range(13,22):
                cell_to_write = ws.cell(row=y, column=j)
                ws.cell(row=y, column=j).border = thin_border
                cell_to_write.value = arr[x]
                x+=1
            i=i+Mod   ##incrementing the value of i to i+Mod
            y=y+1

y=y+3  
dic={}  ##empty dictionary
## adding values in dic 
for i in range(0,4):
    dic[i+1]=2*i+1
    dic[-(i+1)]=2*(i+1)
## 2D arr of 9*9 and initilizing the values with 0
rows, cols = (9,9) 
arr = [[0 for i in range(cols)] for j in range(rows)]

## adding values in the arr and updating its val through iterating in the loop
for iteam in range(0,size-1):
    arr[dic[Octant[iteam]]][dic[Octant[iteam+1]]]+=1

## making table for  Overall Transition Count  
cell_to_write=ws.cell(row=y,column=13)
cell_to_write.value='Overall Transition Count'
y=y+1
cell_to_write=ws.cell(row=y,column=14)
cell_to_write.value='To'
cell_to_write=ws.cell(row=y+2,column=12)
cell_to_write.value='From'
y=y+1

# list
id=[1,-1,2,-2,3,-3,4,-4]
arr[0][0]="count"
##adding values in the arr
for i in range(1,9):
    arr[0][i]=id[i-1]
    arr[i][0]=id[i-1]


for i in range(y,y+9):
    for j in range (13,22):
        cell_to_write=ws.cell(row=i,column=j)
        ws.cell(row=i, column=j).border = thin_border
        cell_to_write.value=arr[i-y][j-13]
## giving space by increasing the row          
y=y+9

        
 ##--> saving the workbook in the output file   
wb.save('output_octant_transition_identify.xlsx')